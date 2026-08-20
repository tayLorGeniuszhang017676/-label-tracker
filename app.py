"""
📦 面单识别 → Tracking 回填工具
文件名 = 序号_订单号 → 下划线后的部分作为订单号 → OCR 识别面单上 TRACKING 后面的号码 → 按订单号匹配回填 Excel
"""

import streamlit as st
import pandas as pd
import re
import io
from pathlib import Path
from PIL import Image, ImageFilter, ImageEnhance
import pytesseract

try:
    from pdf2image import convert_from_bytes
    PDF_SUPPORT = True
except ImportError:
    PDF_SUPPORT = False

try:
    import openpyxl
except ImportError:
    openpyxl = None

# ── Page config ──────────────────────────────────────────────────────────────

st.set_page_config(page_title="面单识别 → Tracking 回填", page_icon="📦", layout="wide")

st.markdown("""
<style>
    div[data-testid="stMetric"] {
        background: #f8fafc; border: 1px solid #e2e8f0;
        border-radius: 8px; padding: 12px 16px;
    }
</style>
""", unsafe_allow_html=True)

# ── Session state ────────────────────────────────────────────────────────────

for key in ["extracted_labels", "match_results"]:
    if key not in st.session_state:
        st.session_state[key] = []
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None

# ── 文件名解析：序号_订单号 ───────────────────────────────────────────────────

def parse_filename(filename: str):
    """文件名格式：序号_订单号.ext
    e.g. '6_114-7801838-6969825.png' → 序号 '6', 订单号 '114-7801838-6969825'
    订单号 = 第一个下划线之后的全部内容（用于匹配 Excel 平台单号）。
    没有下划线时，整个文件名作为订单号。
    """
    stem = Path(filename).stem                      # 去掉 .png / .pdf 等后缀
    stem = re.sub(r'\s*\(\d+\)\s*$', '', stem)      # 去掉重复文件的 " (1)" 等尾巴
    stem = stem.strip()

    if "_" not in stem:
        return "", stem

    seq, order = stem.split("_", 1)
    return seq.strip(), order.strip().lstrip("_").strip()


# ── OCR ──────────────────────────────────────────────────────────────────────

def preprocess_image(img: Image.Image) -> Image.Image:
    img = img.convert("L")
    img = ImageEnhance.Contrast(img).enhance(2.0)
    img = img.filter(ImageFilter.SHARPEN)
    w, h = img.size
    if w < 1000:
        img = img.resize((w * 2, h * 2), Image.LANCZOS)
    return img


def extract_tracking(text: str) -> str:
    """从 OCR 文本提取 Tracking：优先取 TRACKING 标签后面的号码。"""
    # 1) 优先：TRACKING #: 后面同一行的内容（如 'TRACKING #: 1Z 251 D1F 02 0668 8702'）
    match = re.search(r'TRACK(?:ING)?\s*(?:#|NO|NUMBER)?\s*[:：#]?\s*([A-Z0-9][A-Z0-9 ]{8,40})', text, re.IGNORECASE)
    if match:
        clean = re.sub(r'\s+', '', match.group(1)).upper()
        if clean.startswith('1Z'):
            return clean[:18]           # UPS 固定 18 位
        if len(clean) >= 10:
            return clean

    # 2) UPS：正文任意位置的 1Z 号码
    match = re.search(r'(1Z[ A-Z0-9]{16,30})', text, re.IGNORECASE)
    if match:
        return re.sub(r'\s+', '', match.group(1)).upper()[:18]

    # 3) USPS：20-22 位数字
    match = re.search(r'\b(\d{20,22})\b', text)
    if match:
        return match.group(1)

    # 4) FedEx：12-15 位数字
    match = re.search(r'\b(\d{12,15})\b', text)
    if match:
        return match.group(1)

    return ""


def extract_recipient(text: str) -> str:
    """从 OCR 文本提取收件人姓名（仅供参考）。"""
    match = re.search(r'SHIP\s*TO\s*:?\s*\n\s*(.+?)(?:\n|$)', text, re.IGNORECASE)
    if match:
        name = re.sub(r"[^\w\s'-]", '', match.group(1).strip()).strip()
        if len(name) >= 2 and not name.isdigit() and not re.match(r'^\d', name):
            return name
    return ""


def process_file(file_bytes: bytes, filename: str) -> list:
    """处理一个上传文件：文件名 → 订单号；OCR → Tracking。"""
    ext = Path(filename).suffix.lower()
    seq, order_number = parse_filename(filename)
    results = []

    if ext == ".pdf":
        if not PDF_SUPPORT:
            return [{"filename": filename, "seq": seq, "order_number": order_number,
                     "recipient": None, "tracking": None, "ocr_text": "",
                     "error": "PDF 支持未安装"}]
        try:
            images = convert_from_bytes(file_bytes, dpi=300)
        except Exception as e:
            return [{"filename": filename, "seq": seq, "order_number": order_number,
                     "recipient": None, "tracking": None, "ocr_text": "",
                     "error": f"PDF 转换失败: {e}"}]

        for i, img in enumerate(images):
            processed = preprocess_image(img)
            text = pytesseract.image_to_string(processed, config='--psm 6')
            tracking = extract_tracking(text)
            recipient = extract_recipient(text)
            if tracking:  # 只保留识别到 Tracking 的页
                page_label = f"{filename} (p{i+1})" if len(images) > 1 else filename
                results.append({
                    "filename": page_label, "seq": seq,
                    "order_number": order_number,
                    "recipient": recipient, "tracking": tracking,
                    "ocr_text": text, "error": None,
                })

        if not results:
            all_text = ""
            for img in images:
                all_text += pytesseract.image_to_string(preprocess_image(img), config='--psm 6') + "\n"
            tracking = extract_tracking(all_text)
            recipient = extract_recipient(all_text)
            results.append({
                "filename": filename, "seq": seq,
                "order_number": order_number,
                "recipient": recipient, "tracking": tracking,
                "ocr_text": all_text[:500],
                "error": None if tracking else "未识别到 Tracking Number",
            })
    else:
        img = Image.open(io.BytesIO(file_bytes))
        processed = preprocess_image(img)
        text = pytesseract.image_to_string(processed, config='--psm 6')
        tracking = extract_tracking(text)
        recipient = extract_recipient(text)
        results.append({
            "filename": filename, "seq": seq,
            "order_number": order_number,
            "recipient": recipient, "tracking": tracking,
            "ocr_text": text,
            "error": None if tracking else "未识别到 Tracking Number",
        })

    return results


# ── Excel ────────────────────────────────────────────────────────────────────

def read_excel(excel_bytes: bytes):
    """Read Excel and return records with order numbers."""
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    main_sheet = wb.sheetnames[0]
    ws = wb[main_sheet]

    # Find columns
    cols = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if not val:
            continue
        s = str(val)
        if "Platform Number" in s or "平台单号" in s:
            cols["order"] = col
        if "Tracking" in s or "跟踪号" in s:
            cols["tracking"] = col
        if "收件人" in s or "RecipientName" in s:
            cols["recipient"] = col

    if "order" not in cols or "tracking" not in cols:
        return None, None

    records = []
    for row in range(2, ws.max_row + 1):
        order_val = ws.cell(row=row, column=cols["order"]).value
        tracking_val = ws.cell(row=row, column=cols["tracking"]).value
        recipient_val = ws.cell(row=row, column=cols.get("recipient", 1)).value if "recipient" in cols else ""

        if order_val:
            records.append({
                "excel_row": row,
                "订单号": str(order_val).strip(),
                "收件人": str(recipient_val).strip() if recipient_val else "",
                "现有 Tracking": str(tracking_val).strip() if tracking_val else "",
            })

    # If recipient is VLOOKUP with no cached value, try source sheet
    if "recipient" in cols and all(r["收件人"] in ("", "None") for r in records):
        for sn in wb.sheetnames:
            if sn != main_sheet:
                ws2 = wb[sn]
                src_rcpt = src_order = None
                for col in range(1, min(30, ws2.max_column + 1)):
                    val = ws2.cell(row=1, column=col).value
                    if val:
                        if "收件人" in str(val): src_rcpt = col
                        if "订单号" in str(val): src_order = col
                if src_rcpt and src_order:
                    o2n = {}
                    for row in range(2, ws2.max_row + 1):
                        o = ws2.cell(row=row, column=src_order).value
                        n = ws2.cell(row=row, column=src_rcpt).value
                        if o and n:
                            o2n[str(o).strip()] = str(n).strip()
                    for rec in records:
                        if rec["订单号"] in o2n:
                            rec["收件人"] = o2n[rec["订单号"]]
                    break

    return records, {
        "tracking_col": cols["tracking"],
        "main_sheet": main_sheet,
    }


# ── UI ───────────────────────────────────────────────────────────────────────

st.markdown("# 📦 面单识别 → Tracking 回填")
st.markdown("上传面单（**文件名 = 序号_订单号**），订单号取下划线后的部分，Tracking 从面单 OCR 识别，按订单号精准匹配回填到 Excel")
st.caption("✅ 完全免费 | 支持 PDF + 图片批量 | 按订单号精准匹配")

# ── Step 1 ───────────────────────────────────────────────────────────────────

st.markdown("---")
st.markdown("### ① 上传面单")
st.caption("⚠️ 文件名格式：`序号_订单号`，如 `6_114-7801838-6969825.png`（第一个下划线**后面**的部分 = 订单号，用于匹配 Excel；Tracking 从面单图片上识别 TRACKING 后面的号码）")

uploaded_files = st.file_uploader(
    "选择面单文件（PDF / PNG / JPG，可多选）",
    type=["pdf", "png", "jpg", "jpeg", "webp", "bmp"],
    accept_multiple_files=True,
    key="label_uploader",
)

if uploaded_files:
    # Preview file list with extracted order numbers
    preview_data = []
    for f in uploaded_files:
        seq, order = parse_filename(f.name)
        preview_data.append({
            "文件": f.name,
            "序号": seq,
            "提取的订单号": order,
        })
    st.dataframe(pd.DataFrame(preview_data), use_container_width=True, hide_index=True)

    if st.button("🔍 开始识别 Tracking", type="primary", use_container_width=True):
        all_results = []
        progress = st.progress(0, text="正在识别...")

        for i, f in enumerate(uploaded_files):
            progress.progress((i + 1) / len(uploaded_files),
                            text=f"正在处理 {f.name} ({i+1}/{len(uploaded_files)})")
            file_results = process_file(f.getvalue(), f.name)
            for r in file_results:
                all_results.append({
                    "文件名": r["filename"],
                    "序号": r.get("seq") or "",
                    "订单号": r["order_number"],
                    "Tracking #": r.get("tracking") or "",
                    "收件人 (参考)": r.get("recipient") or "",
                    "状态": f"❌ {r['error']}" if r.get("error") else "✅ 成功",
                    "_ocr": r.get("ocr_text", ""),
                })

        progress.empty()
        st.session_state.extracted_labels = all_results
        n_ok = sum(1 for r in all_results if r["状态"].startswith("✅"))
        if n_ok > 0:
            st.success(f"识别完成！成功提取 {n_ok}/{len(all_results)} 条 Tracking")
        else:
            st.error("识别失败，展开下方调试信息查看 OCR 原文")

# Show results
if st.session_state.extracted_labels:
    st.markdown("**识别结果：**")
    df = pd.DataFrame(st.session_state.extracted_labels)
    st.dataframe(df[[c for c in df.columns if not c.startswith("_")]],
                use_container_width=True, hide_index=True)

    with st.expander("🔍 调试：OCR 原始文本"):
        for r in st.session_state.extracted_labels:
            st.markdown(f"**{r['文件名']}:**")
            st.code(r.get("_ocr", ""), language=None)

# ── Step 2 ───────────────────────────────────────────────────────────────────

successful = [l for l in st.session_state.extracted_labels if l["状态"].startswith("✅")]
if successful:
    st.markdown("---")
    st.markdown("### ② 上传 ParcelOutbound Excel")

    uploaded_excel = st.file_uploader("选择 Excel 文件", type=["xlsx", "xls"], key="excel_uploader")

    if uploaded_excel:
        excel_bytes = uploaded_excel.getvalue()
        st.session_state.excel_bytes = excel_bytes
        records, meta = read_excel(excel_bytes)

        if records is None:
            st.error("找不到「平台单号」或「Tracking」列，请确认 Excel 格式")
            st.stop()
        if len(records) == 0:
            st.error("Excel 中没有订单数据")
            st.stop()

        st.session_state.excel_records = records
        st.session_state.meta = meta
        st.success(f"已加载 **{len(records)}** 条订单")

        # Build order number lookup
        order_lookup = {}  # order_number -> index in records
        for idx, rec in enumerate(records):
            order_lookup[rec["订单号"]] = idx

        # ── Step 3: Match ────────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### ③ 匹配结果（按订单号精准匹配）")

        match_results = []
        for label in successful:
            order = label["订单号"]
            if order in order_lookup:
                idx = order_lookup[order]
                rec = records[idx]
                match_results.append({
                    "订单号": order,
                    "Tracking": label["Tracking #"],
                    "Excel 收件人": rec["收件人"],
                    "面单收件人": label.get("收件人 (参考)", ""),
                    "现有 Tracking": rec["现有 Tracking"],
                    "excel_row": rec["excel_row"],
                    "匹配": "✅ 精准匹配",
                    "接受": True,
                })
            else:
                match_results.append({
                    "订单号": order,
                    "Tracking": label["Tracking #"],
                    "Excel 收件人": "",
                    "面单收件人": label.get("收件人 (参考)", ""),
                    "现有 Tracking": "",
                    "excel_row": -1,
                    "匹配": "❌ Excel 中无此订单号",
                    "接受": False,
                })

        st.session_state.match_results = match_results

        n_matched = sum(1 for m in match_results if m["excel_row"] > 0)
        n_miss = len(match_results) - n_matched

        c1, c2, c3 = st.columns(3)
        c1.metric("总面单", len(match_results))
        c2.metric("匹配成功", n_matched)
        c3.metric("未匹配", n_miss)

        st.markdown("**确认要回填的条目：**")
        df_match = pd.DataFrame(match_results)
        edited = st.data_editor(
            df_match[["接受", "订单号", "匹配", "Tracking", "Excel 收件人", "面单收件人", "现有 Tracking"]],
            use_container_width=True, hide_index=True,
            disabled=["订单号", "匹配", "Tracking", "Excel 收件人", "面单收件人", "现有 Tracking"],
            column_config={
                "接受": st.column_config.CheckboxColumn("✓ 接受", default=False),
                "Tracking": st.column_config.TextColumn(width="large"),
            },
        )
        for i, acc in enumerate(edited["接受"].tolist()):
            match_results[i]["接受"] = acc

        # ── Step 4: Download ─────────────────────────────────────────────
        st.markdown("---")
        st.markdown("### ④ 下载更新后的 Excel")
        n_acc = sum(1 for m in match_results if m["接受"] and m["excel_row"] > 0)
        if n_acc == 0:
            st.info("没有可回填的条目")
        else:
            st.markdown(f"将回填 **{n_acc}** 条 Tracking Number")
            if st.button(f"⬇️ 生成更新后的 Excel（{n_acc} 条）", type="primary", use_container_width=True):
                wb_w = openpyxl.load_workbook(io.BytesIO(st.session_state.excel_bytes))
                ws_w = wb_w[st.session_state.meta["main_sheet"]]
                filled = 0
                for m in match_results:
                    if m["接受"] and m["excel_row"] > 0 and m["Tracking"]:
                        ws_w.cell(row=m["excel_row"], column=st.session_state.meta["tracking_col"]).value = m["Tracking"]
                        filled += 1
                out = io.BytesIO()
                wb_w.save(out)
                out.seek(0)
                st.download_button(
                    f"📥 下载 ParcelOutbound_Updated.xlsx（{filled} 条已填）",
                    data=out.getvalue(),
                    file_name="ParcelOutbound_Updated.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
                st.success(f"✅ 共填入 {filled} 条 Tracking Number")

# ── Sidebar ──────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### 📖 使用说明")
    st.markdown("""
    1. 上传面单文件（**文件名 = 序号_订单号**）
    2. 点击「开始识别」：订单号取文件名下划线后的部分，Tracking 从面单 OCR 识别
    3. 上传 ParcelOutbound Excel
    4. 按订单号自动精准匹配
    5. 确认后下载更新的 Excel
    """)
    st.divider()
    st.markdown("### ⚠️ 重要")
    st.markdown("""
    文件名必须是 `序号_订单号`！
    例如：`6_114-7801838-6969825.png`
    第一个下划线**之后**的部分用于匹配
    Excel 中的「Platform Number/平台单号」列；
    Tracking 识别面单上 `TRACKING #:`
    后面的号码（自动去空格）
    """)
    st.divider()
    st.markdown("### ℹ️ 支持格式")
    st.markdown("**面单：** PDF / PNG / JPG")
    st.markdown("**快递：** UPS / FedEx / USPS")
    st.divider()
    st.caption("✅ 完全免费，无需 API Key")
